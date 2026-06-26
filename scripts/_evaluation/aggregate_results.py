"""
Aggregate ALL evaluation results into one master overview
===========================================================
Walks <run>/results/{loso,lodo,cda,mida}/<model_variant>/*.json, parses
every result file (regardless of its protocol-specific JSON structure), and
consolidates everything into one tidy "long format" table — one row per
(method, model, architecture-variant, evaluation-scope) — plus a couple of
quick-scan pivot views for the headline metrics.

This is meant to be the single place where you can see, side by side:
  - all 3 frozen encoders (VideoMAE, DINOv2, V-JEPA)
  - all architecture variants (DINOv2's "default" BiLSTM/linear-probe choice
    PLUS the "opposite-arm" comparison scripts: *_linear / *_bilstm)
  - all 4 evaluation protocols (LOSO, LODO, CDA, MIDA)

─────────────────────────────────────────────────────────────────────────────
WHY ONE SCRIPT FOR FOUR DIFFERENT JSON SHAPES
─────────────────────────────────────────────────────────────────────────────
Each protocol's output script reports its "headline" numbers at a different
natural granularity (this mirrors what each script already prints as its own
summary — we're just consolidating, not re-deriving anything):

  LOSO / MIDA  →  output["datasets"] is a DICT keyed by dataset name; each
                  entry has a "summary" with {metric}_mean / {metric}_std,
                  aggregated over n_subjects LOSO folds.
                  ⇒ one overview row per (dataset), values = fold-aggregated
                    mean ± std.

  LODO         →  output["folds"] is a LIST of per-held-out-dataset results
                  (each held-out dataset = exactly one "fold" in this
                  protocol — there is no further repetition), PLUS a top-level
                  "summary" aggregated over the held-out datasets.
                  ⇒ one row per held-out dataset (single-run values, no
                    std), PLUS one "overall" row (mean ± std over the 4).

  CDA          →  output["pairs"] is a LIST of the 12 source→target results
                  — by design a SINGLE train/test split per pair (see
                  videomae_cda.py's rationale for why no repeated folds).
                  ⇒ one row per source→target pair (single-run values).

The parser auto-detects which shape a file has (by checking whether
"datasets" is a dict vs. a list, whether "pairs" exists, whether "folds"
entries carry "held_out") — so this script keeps working unchanged if you
add more result files later (e.g. videomae_mida.py / vjepa_mida.py runs,
or new "opposite-arm" variants).

─────────────────────────────────────────────────────────────────────────────
MODEL / VARIANT DETECTION
─────────────────────────────────────────────────────────────────────────────
The result directory name encodes both the encoder and the architecture
choice, e.g.:
    .../loso/dinov2_bilstm     → model=dinov2,   variant=default  (BiLSTM = native LOSO probe)
    .../mida/dinov2            → model=dinov2,   variant=default  (ridge = native MIDA probe)
    .../mida/dinov2_bilstm     → model=dinov2,   variant=bilstm   (opposite-arm BiLSTM)
    .../cda/videomae           → model=videomae, variant=default
"default" means "whatever dinov2_loso.py / dinov2_lodo.py / etc. already use
natively for that protocol" (BiLSTM for LOSO, linear probe for LODO/CDA/MIDA
— see those scripts' docstrings) — i.e. NOT "no architecture", but "the
architecture the original pipeline considers the right default for that
protocol". "linear"/"bilstm" mark the explicit "opposite-arm" experiments.

─────────────────────────────────────────────────────────────────────────────
DEDUPLICATION
─────────────────────────────────────────────────────────────────────────────
If multiple result files exist for the same (method, model, variant) — e.g.
re-runs on different days — only the MOST RECENT (by the run timestamp in the
filename) is kept; superseded files are logged but skipped. This keeps the
overview free of duplicate/stale rows while still telling you what was
ignored (in case that's not what you wanted).

Output
------
outputs/runs/<run>/aggregate/
    overview_all_results_{ts}.csv         — the master long-format table
    overview_pivot_{metric}_{ts}.csv      — one quick-scan pivot per headline
                                             metric (rows=scope, columns=
                                             method×model×variant)

Usage
-----
    cd main_project
    python scripts/_evaluation/aggregate_results.py
"""

import json
import logging
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parents[2]
import sys as _sys
_sys.path.insert(0, str(Path(__file__).resolve().parent))
from run_paths import output_root          # honours $BEP_OUTPUT_ROOT (per-run folder)
RESULTS_ROOT = output_root() / "results"
OVERVIEW_DIR = output_root() / "aggregate"

METHODS = ["loso", "lodo", "cda", "mida"]
MODELS  = ["videomae", "dinov2", "vjepa", "vjepa2"]   # canonical ordering for display
VARIANT_ORDER = ["default", "linear", "bilstm"]    # canonical ordering for display

METRICS = ["f1_macro_03", "f1_macro_02", "f1_weighted",
           "accuracy", "mae", "qwk", "spearman"]
HEADLINE_METRICS = ["f1_macro_02", "qwk", "accuracy"]   # get their own pivot CSV

TS_RE = re.compile(r"(\d{8}_\d{6})")
# ─────────────────────────────────────────────────────────────────────────────


def parse_model_variant(dirname: str, method=None):
    """
    'dinov2'        -> ('dinov2',   'default')   [ridge probe, native for non-LOSO]
    'dinov2_linear' -> ('dinov2',   'linear')
    'dinov2_bilstm' -> ('dinov2',   'bilstm')    [but for LOSO the BiLSTM IS the
                                                  native probe -> ('dinov2','default')]
    'videomae'      -> ('videomae', 'default')
    """
    for model in MODELS:
        if dirname == model:
            return model, "default"
        if dirname.startswith(model + "_"):
            suffix = dirname[len(model) + 1:]
            # LOSO's native DINOv2 probe is the BiLSTM → keep it as 'default'.
            if model == "dinov2" and suffix == "bilstm" and method == "loso":
                return model, "default"
            return model, suffix
    return dirname, "default"


def latest_per_group(json_paths):
    """
    Keep only the most recent result file (by filename timestamp) per
    (method, model, variant) group. Returns {key: (path, timestamp)} for the
    keepers, and logs any superseded files that were skipped.
    """
    candidates = {}
    for p in json_paths:
        method = p.parents[1].name                       # .../{method}/{dir}/{file}.json
        model, variant = parse_model_variant(p.parents[0].name, method)
        m = TS_RE.search(p.stem)
        ts = m.group(1) if m else "00000000_000000"
        key = (method, model, variant)
        candidates.setdefault(key, []).append((p, ts))

    keepers = {}
    for key, entries in candidates.items():
        entries.sort(key=lambda e: e[1])     # ascending by timestamp
        keepers[key] = entries[-1]
        for p, ts in entries[:-1]:
            log.info("  (superseded, skipped) %s  [%s]", p.relative_to(PROJECT_ROOT), ts)
    return keepers


def rows_from_loso_or_mida(d, method, model, variant, source_file):
    """output['datasets'] is a DICT: {dataset_name: {..., summary: {...}}}"""
    rows = []
    for ds_name, res in d.get("datasets", {}).items():
        if not isinstance(res, dict) or "summary" not in res:
            continue
        summary = res["summary"]
        row = {
            "method": method, "model": model, "variant": variant,
            "scope_type": "dataset (LOSO-fold-aggregated)", "scope": ds_name,
            "n_subjects": res.get("n_subjects"),
            "n_clips":    res.get("n_clips") or res.get("n_target_clips"),
            "n_folds":    res.get("n_folds"),
            "best_hp":    json.dumps(res["best_hp"]) if res.get("best_hp") else None,
            "source_file": source_file,
        }
        for metric in METRICS:
            row[metric]            = summary.get(f"{metric}_mean")
            row[f"{metric}_std"]   = summary.get(f"{metric}_std")
        rows.append(row)
    return rows


def rows_from_lodo(d, method, model, variant, source_file):
    """output['folds'] is a LIST of per-held-out-dataset results + top-level 'summary'."""
    rows = []
    for fold in d.get("folds", []):
        row = {
            "method": method, "model": model, "variant": variant,
            "scope_type": "held-out dataset (single split)", "scope": fold.get("held_out"),
            "n_subjects": None,
            "n_clips":    fold.get("n_test_clips"),
            "n_folds":    None,
            "best_hp":    None,
            "source_file": source_file,
        }
        for metric in METRICS:
            row[metric]          = fold.get(metric)
            row[f"{metric}_std"] = None
        rows.append(row)

    summary = d.get("summary")
    if summary:
        row = {
            "method": method, "model": model, "variant": variant,
            "scope_type": "overall (mean over held-out datasets)",
            "scope": "ALL (LODO mean)",
            "n_subjects": None, "n_clips": None,
            "n_folds": len(d.get("folds", [])),
            "best_hp": None, "source_file": source_file,
        }
        for metric in METRICS:
            row[metric]          = summary.get(f"{metric}_mean")
            row[f"{metric}_std"] = summary.get(f"{metric}_std")
        rows.append(row)
    return rows


def rows_from_cda(d, method, model, variant, source_file):
    """output['pairs'] is a LIST of per-source→target results (single split, no agg)."""
    rows = []
    for pair in d.get("pairs", []):
        row = {
            "method": method, "model": model, "variant": variant,
            "scope_type": "source→target pair (single split)",
            "scope": f"{pair.get('source')} → {pair.get('target')}",
            "n_subjects": None,
            "n_clips":    pair.get("n_test_clips"),
            "n_folds":    None,
            "best_hp":    None,
            "source_file": source_file,
        }
        for metric in METRICS:
            row[metric]          = pair.get(metric)
            row[f"{metric}_std"] = None
        rows.append(row)
    return rows


def parse_file(path: Path, method, model, variant):
    with open(path) as f:
        d = json.load(f)
    source_file = str(path.relative_to(PROJECT_ROOT))

    datasets_field = d.get("datasets")
    if isinstance(datasets_field, dict) and any(
        isinstance(v, dict) and "summary" in v for v in datasets_field.values()
    ):
        return rows_from_loso_or_mida(d, method, model, variant, source_file)

    if "pairs" in d:
        return rows_from_cda(d, method, model, variant, source_file)

    if d.get("folds") and "held_out" in d["folds"][0]:
        return rows_from_lodo(d, method, model, variant, source_file)

    log.warning("Unrecognised result structure in %s — skipped.", source_file)
    return []


def _categorical_sort(df):
    df["model"]   = pd.Categorical(df["model"],   categories=MODELS,        ordered=True)
    df["variant"] = pd.Categorical(df["variant"], categories=VARIANT_ORDER + sorted(
        set(df["variant"]) - set(VARIANT_ORDER)), ordered=True)
    df["method"]  = pd.Categorical(df["method"],  categories=METHODS,       ordered=True)
    return df.sort_values(["method", "model", "variant", "scope"]).reset_index(drop=True)


def main() -> None:
    json_paths = sorted(
        p for method in METHODS
        for p in (RESULTS_ROOT / method).glob("*/*.json")
    )
    if not json_paths:
        log.error("No result files found under %s — run the evaluation scripts first.",
                  RESULTS_ROOT)
        return

    log.info("Found %d result files under %s", len(json_paths), RESULTS_ROOT)
    keepers = latest_per_group(json_paths)
    log.info("→ %d (method, model, variant) groups kept (most recent run each):", len(keepers))

    all_rows = []
    for (method, model, variant), (path, ts) in sorted(keepers.items()):
        log.info("  %-5s | %-10s | %-8s ← %s  [%s]",
                 method, model, variant, path.name, ts)
        all_rows.extend(parse_file(path, method, model, variant))

    if not all_rows:
        log.error("No rows parsed — nothing to aggregate.")
        return

    df = _categorical_sort(pd.DataFrame(all_rows))

    lead_cols   = ["method", "model", "variant", "scope_type", "scope",
                   "n_subjects", "n_clips", "n_folds", "best_hp"]
    metric_cols = [c for m in METRICS for c in (m, f"{m}_std")]
    df = df[lead_cols + metric_cols + ["source_file"]]

    OVERVIEW_DIR.mkdir(parents=True, exist_ok=True)
    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── 1. CSV (master long table — for programmatic / pandas use) ──────────
    long_path = OVERVIEW_DIR / f"overview_all_results_{run_ts}.csv"
    df.to_csv(long_path, index=False)
    log.info("Master overview (%d rows × %d cols) saved to %s", *df.shape, long_path)

    # ── Quick-scan pivots for headline metrics (long-format companions) ─────
    pivots = {}
    for metric in HEADLINE_METRICS:
        pivot = df.pivot_table(
            index="scope", columns=["method", "model", "variant"],
            values=metric, aggfunc="first", observed=False,
        )
        pivots[metric] = pivot
        pivot_path = OVERVIEW_DIR / f"overview_pivot_{metric}_{run_ts}.csv"
        pivot.to_csv(pivot_path)
        log.info("Pivot view for '%s' (rows=scope, cols=method×model×variant) → %s",
                 metric, pivot_path)

    # ── Build the grouped, human-readable tables ONCE — reused for the
    #    console printout AND the readable .txt export below. ────────────────
    pd.set_option("display.width", 220)
    fmt = lambda v: f"{v:.3f}" if pd.notna(v) else "—"
    display_cols = ["model", "variant", "scope", "scope_type", "n_clips",
                    "f1_macro_02", "accuracy", "qwk", "mae", "spearman"]
    display_fmt = {"f1_macro_02": fmt, "accuracy": fmt, "qwk": fmt, "mae": fmt, "spearman": fmt}

    grouped_blocks = []   # list of (title, rendered_string) — for both console + .txt
    for method in METHODS:
        sub = df[df["method"] == method]
        if sub.empty:
            continue
        title = f"========================= {method.upper()} ========================="
        rendered = sub[display_cols].to_string(index=False, formatters=display_fmt)
        grouped_blocks.append((title, rendered))

    for title, rendered in grouped_blocks:
        log.info("\n%s\n%s", title, rendered)

    # ── 2. Readable plain-text export — open with any text editor, no
    #    spreadsheet app / locale / delimiter issues whatsoever. ─────────────
    txt_path = OVERVIEW_DIR / f"overview_readable_{run_ts}.txt"
    with open(txt_path, "w") as f:
        f.write("DINOv2 / VideoMAE / V-JEPA — full evaluation overview\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Source files aggregated: {len(keepers)} (most recent run per "
                f"method×model×variant group)\n")
        f.write("Columns shown here: model | variant | scope | scope_type | n_clips | "
                "f1_macro_02 | accuracy | qwk | mae | spearman\n")
        f.write("(— = not available / NaN. Full table incl. all 7 metrics, mean±std, "
                "n_subjects/n_folds, best_hp and source_file is in the .csv / .xlsx "
                "next to this file.)\n")
        for title, rendered in grouped_blocks:
            f.write(f"\n{title}\n{rendered}\n")
    log.info("Readable plain-text overview saved to %s", txt_path)

    # ── 3. Excel workbook — opens with correctly split columns regardless of
    #    locale/CSV-delimiter settings (the issue with the plain .csv in
    #    Numbers/Excel: NL locale expects ';' as the list separator, not ',',
    #    so everything lands in column A). One sheet per view, autosized
    #    columns, frozen header row, 3-decimal number format on metric cols. ──
    xlsx_path = OVERVIEW_DIR / f"overview_all_results_{run_ts}.xlsx"
    metric_like_cols = [c for c in df.columns
                        if c.split("_std")[0] in METRICS or c == "n_clips"]
    # Sheets with a flat, single-row header (safe to autosize/format cell-by-cell).
    # The pivot_* sheets have a 3-level MultiIndex column header → openpyxl
    # represents the merged header cells as `MergedCell`s (no .column_letter,
    # no per-column single header row), so we deliberately leave those at
    # openpyxl's defaults rather than fight that structure.
    flat_sheets = {"all_results", *METHODS}
    from openpyxl.utils import get_column_letter
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="all_results", index=False)
            for metric, pivot in pivots.items():
                # sheet names are capped at 31 chars by Excel
                pivot.to_excel(writer, sheet_name=f"pivot_{metric}"[:31])
            for method in METHODS:
                sub = df[df["method"] == method][display_cols]
                if not sub.empty:
                    sub.to_excel(writer, sheet_name=method, index=False)

            wb = writer.book
            for ws in wb.worksheets:
                ws.freeze_panes = "A2"
                if ws.title not in flat_sheets:
                    continue   # pivot_* sheets: leave MultiIndex headers as-is
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    header_cell = ws.cell(row=1, column=col_idx)
                    header = str(header_cell.value or "")
                    body_cells = [ws.cell(row=r, column=col_idx) for r in range(2, ws.max_row + 1)]
                    max_len = max(
                        (len(f"{c.value:.3f}") if isinstance(c.value, float)
                         else len(str(c.value)) for c in body_cells if c.value is not None),
                        default=0,
                    )
                    ws.column_dimensions[col_letter].width = min(max(max_len, len(header)) + 2, 60)
                    if header in metric_like_cols or header.endswith("_std"):
                        for cell in body_cells:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = "0.000"
        log.info("Excel workbook (sheets: all_results, pivot_*, loso/lodo/cda/mida; "
                 "autosized columns, frozen header, 3-decimal formatting) saved to %s",
                 xlsx_path)
    except Exception as e:
        log.warning("Could not write .xlsx workbook (%s) — CSV and .txt are still "
                    "available. Is openpyxl installed? (`pip install openpyxl`)", e)

    log.info("\nDone. Three equivalent overviews were written to %s :\n"
             "  • overview_readable_%s.txt   ← EASIEST: just open in any text editor\n"
             "  • overview_all_results_%s.xlsx ← nicely formatted, opens directly in Excel/Numbers\n"
             "  • overview_all_results_%s.csv  ← for pandas / programmatic use\n"
             "(the .csv looked broken in your spreadsheet app because NL-locale Excel/"
             "Numbers expects ';' as the column separator, not ','  — the .xlsx sidesteps "
             "that entirely.)",
             OVERVIEW_DIR, run_ts, run_ts, run_ts)


if __name__ == "__main__":
    main()
